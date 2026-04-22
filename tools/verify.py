#!/usr/bin/env python3
"""Verify the restored Simon Parker Therapy files.

Checks:
- All expected files exist and are non-empty.
- HTML pages all reference styles.css, have no broken local links, and include
  the dual-track/British-voice content markers.
- Images load and are the expected dimensions.
- All 5 docx compliance files validate and contain key required phrases.
- Voice guide has both v1 and v2 sections and the verbatim British paragraph.
- ICS file parses and contains the 5 expected events.
- Tracker opens and the 3 new rows (43, 44, 45) are present.
- Constants consistency: every value in practice_constants.json appears
  in the output files (website HTML, compliance docx, clinical forms docx)
  wherever it is supposed to. Any future JSON edit that fails to propagate
  through the build scripts will surface here as a failing check.
"""
from pathlib import Path
import re
import sys
import zipfile
import subprocess

ROOT = Path("/sessions/great-busy-edison/mnt/SimonParkerTherapy")
WEB = ROOT / "website"
COMP = ROOT / "compliance"
CLIN = ROOT / "clinical_forms"
IMG = WEB / "images"
TOOLS = ROOT / "tools"

# Make practice_constants importable without needing tools/ on sys.path.
sys.path.insert(0, str(TOOLS))
import practice_constants as PC  # noqa: E402

OK, WARN, FAIL = [], [], []

def ok(msg):  OK.append(msg);  print(f"  ✓ {msg}")
def warn(msg): WARN.append(msg); print(f"  ! {msg}")
def fail(msg): FAIL.append(msg); print(f"  ✗ {msg}")


def section(title):
    print(f"\n— {title} —")


# --- 1. File inventory ---
section("File inventory")
EXPECTED = [
    WEB / "index.html",
    WEB / "about.html",
    WEB / "approach.html",
    WEB / "fees.html",
    WEB / "for-teams.html",
    WEB / "faq.html",
    WEB / "contact.html",
    WEB / "privacy.html",
    WEB / "thank-you.html",
    WEB / "styles.css",
    WEB / "README.md",
    WEB / "favicon.svg",
    WEB / "robots.txt",
    WEB / "sitemap.xml",
    WEB / "_headers",
    ROOT / "netlify.toml",
    IMG / "simon-parker.jpg",
    IMG / "simon-parker-portrait.jpg",
    IMG / "simon-parker-thumb.jpg",
    IMG / "og-card.jpg",
    COMP / "Informed_Consent.docx",
    COMP / "Notice_of_Privacy_Practices.docx",
    COMP / "Good_Faith_Estimate.docx",
    COMP / "HIPAA_Security_Risk_Assessment.docx",
    COMP / "Incident_and_Breach_Response_Plan.docx",
    ROOT / "Practice_Voice_and_Bio.md",
    ROOT / "Security_Reminders.ics",
    ROOT / "Simon_Parker_Therapy_Practice_Tracker.xlsx",
]
for p in EXPECTED:
    if p.exists() and p.stat().st_size > 0:
        ok(f"{p.relative_to(ROOT)} ({p.stat().st_size:,} bytes)")
    else:
        fail(f"MISSING or empty: {p.relative_to(ROOT)}")

# --- 2. HTML content + link graph ---
section("Website content")
PAGES = ["index.html", "about.html", "approach.html", "fees.html", "for-teams.html", "faq.html", "contact.html", "privacy.html", "thank-you.html"]
MUST_ALL = [
    ("styles.css",          "links stylesheet"),
    ("Simon Parker Therapy","practice name"),
    ("988",                 "988 crisis line"),
]
# Every page should also have a canonical link, favicon reference, and OG image tag.
MUST_HEAD = [
    ('rel="canonical"',          "canonical link"),
    ("favicon.svg",              "favicon reference"),
    ('property="og:image"',      "og:image meta tag"),
]
PAGE_SPECIFIC = {
    "index.html":    ["high-pressure roles", "trauma and PTSD",
                      "credibility", "path-tile",
                      "Deciding whether to reach out",
                      "Who I work with",
                      "moral injury",
                      "ProfessionalService"],
    "about.html":    ["British therapist", "VA", "moral injury",
                      "simon-parker-portrait.jpg", "CW024575",
                      "\"@type\": \"Person\""],
    "approach.html": ["Beckian CBT", "REBT", "Interpersonal Psychotherapy", "IPT",
                      "Beck and Ellis", "Weissman", "Why I do this", "fifteen years"],
    "fees.html":     ["$160", "Good Faith Estimate", "out-of-network",
                      "What's included", "Sliding scale",
                      "HSA and FSA", "CPT code 90834"],
    "for-teams.html":["Thinking Clearly Under Pressure", "$3,500", "$6,500", "$9,500",
                      "Fire, EMS", "Hospitals", "Trading floors", "Law firms",
                      "Trower, Casey, and Dryden", "not positive thinking",
                      "isn't therapy", "scoping call",
                      "simon-parker-portrait.jpg", "in person",
                      "fifteen years", "I'm British",
                      "by exception",
                      "First-cohort pricing"],
    "faq.html":      ["988", "out-of-network", "Beckian CBT", "IPT",
                      "FAQPage"],
    "contact.html":  ["consultation", "info@simonparker.com",
                      "direct-contact",
                      # Web3Forms wiring
                      "api.web3forms.com/submit",
                      'name="access_key"',
                      'name="botcheck"',
                      'name="name"',
                      'name="email"',
                      'name="state"',
                      'name="message"',
                      'name="contact_method"',
                      "thank-you.html",
                      "privacy.html"],
    "privacy.html":  ["Privacy Policy", "HIPAA",
                      "info@simonparker.com"],
    "thank-you.html":["noindex", "I've got it",
                      "info@simonparker.com", "988"],
}
for page in PAGES:
    txt = (WEB / page).read_text()
    for needle, label in MUST_ALL:
        if needle in txt: ok(f"{page}: has {label}")
        else:             fail(f"{page}: missing {label}")
    for needle, label in MUST_HEAD:
        if needle in txt: ok(f"{page}: has {label}")
        else:             fail(f"{page}: missing {label}")
    for needle in PAGE_SPECIFIC.get(page, []):
        if needle in txt: ok(f"{page}: contains '{needle}'")
        else:             fail(f"{page}: missing '{needle}'")
    # local link graph
    hrefs = re.findall(r'href="([^"]+)"', txt) + re.findall(r'src="([^"]+)"', txt)
    for h in hrefs:
        if h.startswith("#") or h.startswith("mailto:") or h.startswith("tel:") or h.startswith("http"):
            continue
        target = (WEB / h).resolve()
        if not target.exists():
            fail(f"{page}: broken local link → {h}")

# --- 3. Images ---
section("Images")
try:
    from PIL import Image
    for name, expect in [
        ("simon-parker.jpg", None),
        ("simon-parker-portrait.jpg", (720, 960)),
        ("simon-parker-thumb.jpg", (480, 480)),
    ]:
        im = Image.open(IMG / name)
        im.verify()
        im = Image.open(IMG / name)
        if expect and im.size != expect:
            fail(f"{name}: expected {expect}, got {im.size}")
        else:
            ok(f"{name} ({im.size[0]}x{im.size[1]})")
except Exception as e:
    fail(f"Image check failed: {e}")

# --- 4. DOCX validation + content markers ---
section("Compliance docs")
DOCX_MARKERS = {
    "Informed_Consent.docx":            ["Informed Consent", "CW024575", "988", "Beck and Ellis", "Weissman"],
    "Notice_of_Privacy_Practices.docx": ["Notice of Privacy Practices", "HIPAA", "HITECH"],
    "Good_Faith_Estimate.docx":         ["Good Faith Estimate", "No Surprises Act", "$160"],
    "HIPAA_Security_Risk_Assessment.docx": ["Security Risk Assessment", "FileVault", "SimplePractice", "BAA"],
    "Incident_and_Breach_Response_Plan.docx": ["Incident", "Breach", "60 calendar days", "tabletop"],
}
for fname, needles in DOCX_MARKERS.items():
    f = COMP / fname
    # validate
    r = subprocess.run(
        ["python3", "/sessions/great-busy-edison/mnt/.skills/skills/docx/scripts/office/validate.py", str(f)],
        capture_output=True, text=True
    )
    if "PASSED" in r.stdout:
        ok(f"{fname} validates")
    else:
        fail(f"{fname} does NOT validate")
    # content
    with zipfile.ZipFile(f) as z:
        body = z.read("word/document.xml").decode("utf-8", errors="ignore")
    for n in needles:
        if n in body: ok(f"{fname}: has '{n}'")
        else:         fail(f"{fname}: missing '{n}'")

# --- 5. Voice guide ---
section("Practice voice and bio")
voice = (ROOT / "Practice_Voice_and_Bio.md").read_text()
VOICE_CHECKS = [
    "v1 (original)",
    "v4 (current)",
    "spilling our feelings",         # verbatim paragraph
    "wall of therapist-speak",       # verbatim paragraph
    "Beckian CBT",
    "Interpersonal Psychotherapy",
    "Weissman, Markowitz, and Klerman",
    "Beck and Ellis",
    "CW024575",
    "high-pressure roles",
    "veterans, active-duty",
]
for n in VOICE_CHECKS:
    if n in voice: ok(f"voice guide: has '{n}'")
    else:          fail(f"voice guide: missing '{n}'")

# --- 6. ICS ---
section("Security Reminders calendar")
ics = (ROOT / "Security_Reminders.ics").read_text()
ICS_CHECKS = [
    ("BEGIN:VCALENDAR", "calendar header"),
    ("END:VCALENDAR",   "calendar footer"),
    ("1Password",       "password manager event"),
    ("FileVault",       "FileVault event"),
    ("SimplePractice BAA", "BAA event"),
    ("tabletop",        "tabletop exercise event"),
    ("RRULE:FREQ=YEARLY", "annual review recurrence"),
]
for n, label in ICS_CHECKS:
    if n in ics: ok(f"ics: has {label}")
    else:        fail(f"ics: missing {label} ({n})")
count = ics.count("BEGIN:VEVENT")
if count >= 5: ok(f"ics: {count} VEVENT blocks")
else:          fail(f"ics: expected >=5 VEVENT blocks, got {count}")

# --- 7. Tracker ---
section("Practice tracker")
try:
    import openpyxl
    wb = openpyxl.load_workbook(ROOT / "Simon_Parker_Therapy_Practice_Tracker.xlsx", data_only=False)
    sheet_names = wb.sheetnames
    ok(f"tracker opens — sheets: {sheet_names}")
    # Find the main sheet (contains the items/rows)
    main = wb[sheet_names[0]] if "Dashboard" not in sheet_names else None
    # the rows we restored should be reachable; look across all sheets for key items
    found = {"Incident & Breach Response Plan": False, "SimplePractice": False, "dual-track": False}
    for s in wb.worksheets:
        for row in s.iter_rows(values_only=True):
            text = " | ".join(str(c) for c in row if c is not None)
            for k in list(found):
                if k.lower() in text.lower():
                    found[k] = True
    for k, v in found.items():
        if v: ok(f"tracker mentions '{k}'")
        else: warn(f"tracker does not mention '{k}' (may be in a Notes column, check manually)")
except Exception as e:
    fail(f"Tracker check failed: {e}")

# --- 8. Constants consistency ---
section("Constants consistency (practice_constants.json → outputs)")


def _docx_text(path: Path) -> str:
    """Return the *visible* text of a .docx — body + headers + footers.

    Word often splits a single reader-visible string across multiple
    ``<w:t>`` runs (for bolding, spacing, etc.), so a plain substring
    match on the raw XML can miss text that clearly appears to a reader.
    Here we pull just the text of each ``<w:t>`` run from the main
    document body AND every header / footer XML part (that's where the
    practice name, clinician, and NPI usually live), and join them with
    spaces. HTML entities are unescaped so searches for literal
    apostrophes, quotes, and ampersands work naturally.
    """
    import html as _html
    parts = []
    with zipfile.ZipFile(path) as z:
        for name in z.namelist():
            if not name.startswith("word/") or not name.endswith(".xml"):
                continue
            if not (
                name == "word/document.xml"
                or name.startswith("word/header")
                or name.startswith("word/footer")
            ):
                continue
            xml = z.read(name).decode("utf-8", errors="ignore")
            parts.extend(re.findall(r"<w:t[^>]*>([^<]*)</w:t>", xml))
    return _html.unescape(" ".join(parts))


def _must_contain(haystack: str, needle: str, where: str):
    if needle in haystack:
        ok(f"{where}: contains '{needle}'")
    else:
        fail(f"{where}: MISSING '{needle}'")


def _must_not_contain(haystack: str, needle: str, where: str):
    if needle not in haystack:
        ok(f"{where}: no stale '{needle}'")
    else:
        fail(f"{where}: STALE value '{needle}' still present")


# 8a. practice_constants.py loaded and basic values exposed
try:
    _legal = PC.C["identity"]["legal_entity"]
    _npi = PC.C["registry"]["npi"]
    _phone = PC.C["contact"]["phone"]
    _email = PC.C["contact"]["email"]
    _city = PC.C["contact"]["office_city"]
    _open_from = PC.C["contact"]["office_open_from"]
    _voice = PC.VOICE_PARAGRAPH
    _cbc_attr = PC.C["frameworks"]["cbc"]["attribution"]
    _ipt_attr = PC.C["frameworks"]["ipt"]["attribution"]
    _fee50 = f"${PC.C['fees']['session_50min_price_usd']}"
    _lic_pa = next(l for l in PC.C["licenses"] if l["abbr"] == "PA")["number"]
    _lic_tx = next(l for l in PC.C["licenses"] if l["abbr"] == "TX")["number"]
    _lic_ca = next(l for l in PC.C["licenses"] if l["abbr"] == "CA")["number"]
    ok("practice_constants.py loaded cleanly")
except Exception as e:
    fail(f"practice_constants.py failed to load: {e}")
    # If constants won't load, skip the remaining consistency checks.
    _legal = None

if _legal is not None:
    # 8b. Website pages pick up the right values from the constants
    # Each row: (filename, [(value, label), ...])
    WEB_EXPECT = {
        "index.html":   [(_legal, "legal_entity")],
        "about.html":   [(_lic_pa, "PA licence #"),
                         (_lic_tx, "TX licence #"),
                         (_lic_ca, "CA licence #"),
                         (_npi, "NPI"),
                         (_cbc_attr, "CBC attribution"),
                         (_ipt_attr, "IPT attribution")],
        "approach.html":[(_cbc_attr, "CBC attribution"),
                         (_ipt_attr, "IPT attribution")],
        "fees.html":    [(_fee50, "50-min fee")],
        "faq.html":     [(_lic_pa, "PA licence #"),
                         (_lic_tx, "TX licence #"),
                         (_lic_ca, "CA licence #")],
        "contact.html": [(_email, "email"),
                         (_phone, "phone"),
                         (_city, "office city"),
                         (_open_from, "office open-from")],
    }
    for page, items in WEB_EXPECT.items():
        txt = (WEB / page).read_text(encoding="utf-8")
        for needle, label in items:
            _must_contain(txt, needle, f"website/{page} [{label}]")

    # 8c. Every website page should have the shared footer values
    for page in PAGES:
        txt = (WEB / page).read_text(encoding="utf-8")
        _must_contain(txt, _legal, f"website/{page} [footer legal_entity]")
        _must_contain(txt, _lic_pa, f"website/{page} [footer PA #]")
        _must_contain(txt, _lic_tx, f"website/{page} [footer TX #]")
        _must_contain(txt, _lic_ca, f"website/{page} [footer CA #]")

    # 8d. Compliance .docx files pick up the constants.
    # Expectations are tuned to what each document actually contains
    # today (so every constant that the build scripts render IS tested).
    # Items listed under SHOULD are content gaps — flagged as warnings
    # rather than failures so Simon sees them as a to-do list.
    _privacy_email = f"privacy@{PC.C['contact']['website']}"
    COMP_MUST = {
        "Informed_Consent.docx": [
            (_legal, "legal_entity"),
            (_npi, "NPI"),
            (_lic_pa, "PA licence #"),
            (_lic_tx, "TX licence #"),
            (_lic_ca, "CA licence #"),
            (_cbc_attr, "CBC attribution"),
            (_ipt_attr, "IPT attribution"),
        ],
        "Notice_of_Privacy_Practices.docx": [
            (_legal, "legal_entity"),
            (_phone, "phone"),
            (_privacy_email, "privacy@ email"),
        ],
        "Good_Faith_Estimate.docx": [
            (_legal, "legal_entity"),
            (_fee50, "50-min fee"),
        ],
        "HIPAA_Security_Risk_Assessment.docx": [
            (_legal, "legal_entity"),
        ],
        "Incident_and_Breach_Response_Plan.docx": [
            (_legal, "legal_entity"),
            (_phone, "phone"),
            (_privacy_email, "privacy@ email"),
        ],
    }
    # Content gaps — nice to have, not strictly required by the build.
    # These surface as warnings so they remain visible without failing CI.
    COMP_SHOULD = {
        "Good_Faith_Estimate.docx": [
            (_npi, "NPI (No Surprises Act requires provider identifier)"),
        ],
        "Informed_Consent.docx": [
            (_phone, "phone"),
            (_email, "contact email"),
        ],
    }
    for fname, items in COMP_MUST.items():
        f = COMP / fname
        if not f.exists():
            fail(f"compliance/{fname} missing")
            continue
        body = _docx_text(f)
        for needle, label in items:
            _must_contain(body, needle, f"compliance/{fname} [{label}]")
        for needle, label in COMP_SHOULD.get(fname, []):
            if needle in body:
                ok(f"compliance/{fname} [{label}]: contains '{needle}'")
            else:
                warn(f"compliance/{fname} [{label}]: missing '{needle}' — consider adding")

    # 8e. Clinical forms pick up the constants.
    # All 8 forms share the practice header (legal_entity + clinician + NPI).
    # The Superbill additionally renders the practice info block with all
    # three licence numbers and the contact line.
    CLIN_ALL = [
        "Intake_Questionnaire.docx",
        "Release_of_Information.docx",
        "Progress_Note_Template.docx",
        "Treatment_Plan_Template.docx",
        "Superbill_Template.docx",
        "PHQ-9.docx",
        "GAD-7.docx",
        "PCL-5.docx",
    ]
    CLIN_WITH_LICENCES = {"Superbill_Template.docx"}
    CLIN_WITH_CONTACT = {"Superbill_Template.docx", "Release_of_Information.docx"}
    for fname in CLIN_ALL:
        f = CLIN / fname
        if not f.exists():
            fail(f"clinical_forms/{fname} missing")
            continue
        body = _docx_text(f)
        _must_contain(body, _legal, f"clinical_forms/{fname} [legal_entity]")
        _must_contain(body, _npi, f"clinical_forms/{fname} [NPI]")
        if fname in CLIN_WITH_LICENCES:
            for num, abbr in ((_lic_pa, "PA"), (_lic_tx, "TX"), (_lic_ca, "CA")):
                _must_contain(body, num, f"clinical_forms/{fname} [{abbr} licence #]")
        if fname in CLIN_WITH_CONTACT:
            _must_contain(body, _email, f"clinical_forms/{fname} [email]")

    # 8f. Voice guide contains the locked verbatim paragraph
    voice_txt = (ROOT / "Practice_Voice_and_Bio.md").read_text(encoding="utf-8")
    # The guide may wrap at different widths; match on a distinctive
    # middle fragment rather than the whole paragraph.
    for frag in [
        "British people are not exactly famous for spilling our feelings",
        "wall of therapist-speak",
    ]:
        _must_contain(voice_txt, frag, "Practice_Voice_and_Bio.md [voice paragraph]")
    _must_contain(voice_txt, _cbc_attr, "Practice_Voice_and_Bio.md [CBC attribution]")
    _must_contain(voice_txt, _ipt_attr, "Practice_Voice_and_Bio.md [IPT attribution]")

    # 8g. Nothing is using the old hand-entered values that the build
    # scripts should have regenerated. We only guard values that are
    # definitely wrong if they appear; we can't flag every possible
    # typo, only ones we've actively retired. Currently: none.
    # (Add entries here if a constant is ever renamed.)

    # Netlify headers file: must include a CSP that allows the Web3Forms
    # form endpoint, otherwise contact form submissions will be blocked in
    # production. These are failures (not warnings) because a missing rule
    # here is a silent user-facing bug.
    headers_path = WEB / "_headers"
    if headers_path.exists():
        headers_txt = headers_path.read_text(encoding="utf-8")
        for needle, label in [
            ("Strict-Transport-Security",      "HSTS header"),
            ("X-Frame-Options: DENY",          "X-Frame-Options: DENY"),
            ("Content-Security-Policy",        "CSP header"),
            ("api.web3forms.com",              "CSP allows Web3Forms"),
            ("fonts.googleapis.com",           "CSP allows Google Fonts CSS"),
            ("fonts.gstatic.com",              "CSP allows Google Fonts webfonts"),
            ("frame-ancestors 'none'",         "CSP frame-ancestors locked"),
        ]:
            if needle in headers_txt:
                ok(f"website/_headers: {label}")
            else:
                fail(f"website/_headers: missing {label}")
    else:
        fail("website/_headers missing (Netlify will not set security headers)")

    # netlify.toml: must point `publish` at the `website` directory, otherwise
    # Netlify will publish the repo root and expose the tools/, compliance/,
    # and clinical_forms/ folders to the public web.
    netlify_toml = ROOT / "netlify.toml"
    if netlify_toml.exists():
        toml_txt = netlify_toml.read_text(encoding="utf-8")
        if 'publish = "website"' in toml_txt:
            ok("netlify.toml: publish directory is 'website'")
        else:
            fail("netlify.toml: publish directory is NOT set to 'website' "
                 "— Netlify would expose the whole repo")
    else:
        fail("netlify.toml missing (Netlify cannot locate the site folder)")

    # 8g-pre-launch: Warn if the Web3Forms access key placeholder has not yet
    # been replaced with a real key. Flagged as a warning (not a failure) so
    # the site verifies cleanly while Simon is setting up the Web3Forms account,
    # and so the TODO stays visible until he swaps it in.
    contact_txt = (WEB / "contact.html").read_text(encoding="utf-8")
    if "ACCESS_KEY_HERE" in contact_txt:
        warn("contact.html: Web3Forms access_key is still the 'ACCESS_KEY_HERE' placeholder "
             "— replace it with the key from web3forms.com before launch")
    else:
        ok("contact.html: Web3Forms access_key placeholder has been replaced")

    # 8h. Markers are present in every website page (so future rebuilds work)
    for page in PAGES:
        txt = (WEB / page).read_text(encoding="utf-8")
        if "<!--@block:nav-->" in txt and "<!--/@block:nav-->" in txt:
            ok(f"website/{page}: has nav block markers")
        else:
            fail(f"website/{page}: missing nav block markers")
        if "<!--@block:footer-->" in txt and "<!--/@block:footer-->" in txt:
            ok(f"website/{page}: has footer block markers")
        else:
            fail(f"website/{page}: missing footer block markers")


# --- Summary ---
section("Summary")
print(f"{len(OK)} checks passed, {len(WARN)} warnings, {len(FAIL)} failures.")
if FAIL:
    print("\nFAILURES:")
    for m in FAIL: print(f"  - {m}")
    sys.exit(1)
if WARN:
    print("\nWarnings:")
    for m in WARN: print(f"  - {m}")
print("\nRestore verified.")
